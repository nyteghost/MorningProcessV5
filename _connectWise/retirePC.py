from mpConfigs.doorKey import config, cwlogin
import requests
import json
import time
import getpass
from datetime import datetime
import pandas as pd
import getpass
import glob
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from requests.adapters import HTTPAdapter, Retry

retry_strategy = Retry(
    total=10,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["HEAD", "GET", "OPTIONS"]
)
adapter = HTTPAdapter(max_retries=retry_strategy)
http = requests.Session()
http.mount("https://", adapter)
http.mount("http://", adapter)


# Config
tokenHeader = config['cwaHeader']
cwURL = config['cwAPI']['web']
cwAURL = 'https://sca-atl.hostedrmm.com/cwa/api/v1/'

# File Information
localuser = getpass.getuser()
prefix = fr'C:\Users\{localuser}'
excelFolder = r'\Southeastern Computer Associates, LLC\GCA Deployment - Documents\Database\Automate Audit Win10L Returns'
comboBreaker = prefix + excelFolder
my_Date = time.strftime("%Y%m%d")


def getToken():
    loginCRED = cwlogin()
    api_request = cwAURL + '/' + 'apitoken'
    response = requests.post(url=api_request, headers=tokenHeader, json=loginCRED)
    mydict = response.json()
    accessToken = mydict.get('AccessToken')
    data = [accessToken]
    print('Token Generated.')
    if response.status_code != 200:
        print(api_request)
        print(response.status_code)
        print(response.text)
        pass
    return accessToken


def getcwaHEADER(Token):
    token = "Bearer " + Token
    GetHeader = {
        "Authorization": token,
        'clientId': config['cwaHeader']['clientID'],
        "Content-Type": "application/json"
    }
    return GetHeader


def getSpecificComputer(computerName, Token):
    getHeader = getcwaHEADER(Token)
    api_request = cwAURL + '/' + 'Computers?condition=ComputerName = "GCA-{computer}"'.format(computer=computerName)
    response = http.get(url=api_request, headers=getHeader)
    rt = response.text
    try:
        res = json.loads(rt)
    except Exception as e:
        print(e)
        getToken()
        time.sleep(5)
        res = json.loads(rt)
    for i in res:
        # print(i)
        compId = i['Id']
        locId = i['Location']['Id']
        computerName = i['ComputerName']
        serialNumber = i['SerialNumber']
        lastUser = i['LastUserName']
        lastContact = i['RemoteAgentLastContact']
        status = i['Status']

        print(computerName)
        print(compId)
        print(locId)
        print(serialNumber)
        print(lastContact)
        print(status)
        print(lastUser)
        return compId, serialNumber


def retireThatPC(compId, Token):
    my_date = datetime.now()
    getHeader = getcwaHEADER(Token)
    api_request = cwAURL + '/' + 'batch/ScriptExecute'
    my_date = str(my_date)
    payload = {
        "EntityType": 1,
        "EntityIds": [compId],
        "ScriptId": "6431",
        "Schedule": {"ScriptScheduleFrequency": {"ScriptScheduleFrequencyId": 1}},
        "Parameters": [],
        "UseAgentTime": False,
        "StartDate": my_date,
        "OfflineActionFlags": {"SkipsOfflineAgents": False},
        "Priority": 12
    }
    print(payload)
    response = requests.post(url=api_request, headers=getHeader, json=payload)
    rt = response.text
    print(rt)
    print("####################################################################")


def runIt(dictionary):
    complete = []
    not_complete = []
    not_found = []
    authToken = getToken()

    for assetID, serialNum in dictionary.items():
        try:
            computerId, serialNumber = getSpecificComputer(assetID, authToken)
        except TypeError:
            print(f"{assetID} {serialNum} not found in Automate.")
            not_found.append(assetID)
        else:
            if serialNum.strip().lower() == serialNumber.strip().lower():
                time.sleep(2)
                retireThatPC(computerId, authToken)
                complete.append(assetID)
                time.sleep(2)
            else:
                print(f"For asset {assetID}, excel SN: {serialNum} does not match automate: {serialNumber}")
                not_complete.append(assetID)
                time.sleep(2)
    return complete, not_complete, not_found


def colorThoseRows(file, listing):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    fill_gen = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")

    for row in ws.iter_rows():
        if row[0].value in listing:
            for i in range(5):
                row[i].fill = fill_gen

    wb.save(file)
    wb.close()


def main():
    for file in glob.glob(comboBreaker + '\\' + f'*{my_Date}*', recursive=True):
        df = pd.read_excel(file)
        print(df)

        my_dict = df.set_index('AssetID').to_dict()['Serial#']
        print("####################################################################")
        complete, not_complete, not_found = runIt(my_dict)
        colorThoseRows(file, complete)

        print('Found in Automate')
        for i in complete:
            print(i)
        print()
        print('Serials did not match')
        for i in not_complete:
            print(i)
        print()
        print('Not Found in Automate')
        for i in not_found:
            print(i)


main()
